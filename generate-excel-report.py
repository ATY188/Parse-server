#!/usr/bin/env python3
"""
從批次測試結果生成 Excel 報告
可以處理新舊兩種 JSON 格式
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def generate_excel_report(json_file: str, output_file: str = None):
    """
    從 JSON 結果生成 Excel 報告
    
    Args:
        json_file: 輸入的 JSON 檔案路徑
        output_file: 輸出的 Excel 檔案名稱（可選）
    """
    # 載入結果
    with open(json_file, 'r', encoding='utf-8') as f:
        results = json.load(f)
    
    # 建立 Excel
    wb = Workbook()
    
    # ========== MASTER 分頁 ==========
    ws_master = wb.active
    ws_master.title = "詳細資料 (Master)"
    
    # 定義表頭
    headers_master = [
        "編號",
        "連結",
        "解析文字（前500字）",
        "爬蟲類型",
        "耗時（秒）",
        "錯誤原因",
        "標題",
        "作者",
        "發布日期",
        "字數",
        "中文字數",
        "路由決策",
        "嘗試次數",
        "HTTP狀態碼",
        "解析成功"
    ]
    
    # 設定表頭樣式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 寫入表頭
    for col_num, header in enumerate(headers_master, 1):
        cell = ws_master.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填充資料
    for idx, result in enumerate(results, start=2):
        url = result.get('url', '')
        id_num = result.get('id', '???')
        success = result.get('success', False)
        parsed_data = result.get('parsed_data', {})
        error = result.get('error', '')
        elapsed_time = result.get('elapsed_time', 0)
        status_code = result.get('status_code', '')
        routing_decision = result.get('routing_decision', 'N/A')
        attempts = result.get('attempts', 1)
        
        # 處理中文字數統計（防止 None）
        text_content = parsed_data.get('text_content') or ''
        total_chars = len(text_content)
        chinese_chars = len([c for c in text_content if '\u4e00' <= c <= '\u9fff'])
        
        # 爬蟲類型（中英文）
        method = result.get('rendering_method') or parsed_data.get('rendering_method', 'unknown')
        method_display = {
            'static': '靜態爬蟲 (Trafilatura)',
            'playwright': '動態爬蟲 (Playwright)',
            'unknown': '未知'
        }.get(method, method)
        
        # 填充每一列
        row_data = [
            id_num,                                      # 編號
            url,                                         # 連結
            text_content[:500] if text_content else '',  # 解析文字（前500字）
            method_display,                              # 爬蟲類型
            round(elapsed_time, 2) if elapsed_time else 0,  # 耗時（秒）
            error if not success else '',                # 錯誤原因
            parsed_data.get('title') or '',              # 標題
            parsed_data.get('author') or '',             # 作者
            parsed_data.get('date_published') or '',     # 發布日期
            total_chars,                                 # 字數
            chinese_chars,                               # 中文字數
            routing_decision,                            # 路由決策
            attempts,                                    # 嘗試次數
            status_code,                                 # HTTP狀態碼
            '是' if success else '否'                    # 解析成功
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws_master.cell(row=idx, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            
            # 失敗項目標紅色
            if not success:
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # 調整列寬
    column_widths = [10, 60, 70, 30, 12, 40, 40, 20, 15, 10, 12, 15, 12, 15, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws_master.column_dimensions[get_column_letter(col_num)].width = width
    
    # 凍結首列
    ws_master.freeze_panes = "A2"
    
    # ========== SUMMARY 分頁 ==========
    ws_summary = wb.create_sheet(title="測試總結 (Summary)")
    
    # 統計資料
    total = len(results)
    success_count = sum(1 for r in results if r.get('success'))
    fail_count = total - success_count
    success_rate = (success_count / total * 100) if total > 0 else 0
    
    static_count = sum(1 for r in results if (r.get('rendering_method') or r.get('parsed_data', {}).get('rendering_method')) == 'static')
    dynamic_count = sum(1 for r in results if (r.get('rendering_method') or r.get('parsed_data', {}).get('rendering_method')) == 'playwright')
    
    avg_time = sum(r.get('elapsed_time', 0) for r in results) / total if total > 0 else 0
    max_time = max((r.get('elapsed_time', 0) for r in results), default=0)
    min_time = min((r.get('elapsed_time', 0) for r in results if r.get('success')), default=0)
    
    avg_chars = sum(len(r.get('parsed_data', {}).get('text_content') or '') for r in results if r.get('success')) / success_count if success_count > 0 else 0
    
    # 標題
    title_cell = ws_summary.cell(row=1, column=1)
    title_cell.value = f"📊 {total} 個連結批次測試總結報告"
    title_cell.font = Font(bold=True, size=16, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.merge_cells('A1:C1')
    
    # 測試資訊
    ws_summary.cell(row=2, column=1).value = "測試時間："
    ws_summary.cell(row=2, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary.cell(row=3, column=1).value = "資料來源："
    ws_summary.cell(row=3, column=2).value = os.path.basename(json_file)
    
    # 空行
    ws_summary.cell(row=4, column=1)
    
    # 主要統計
    summary_data = [
        ["📈 整體統計", "數值", "百分比/說明"],
        ["總連結數", total, ""],
        ["✅ 成功數", success_count, f"{success_rate:.1f}%"],
        ["❌ 失敗數", fail_count, f"{100-success_rate:.1f}%"],
        ["", "", ""],
        ["🔧 爬蟲方法", "數量", "百分比"],
        ["靜態爬蟲 (Trafilatura)", static_count, f"{static_count/total*100:.1f}%" if total > 0 else "0%"],
        ["動態爬蟲 (Playwright)", dynamic_count, f"{dynamic_count/total*100:.1f}%" if total > 0 else "0%"],
        ["", "", ""],
        ["⏱️ 效能統計", "數值", "單位"],
        ["平均耗時", round(avg_time, 2), "秒"],
        ["最長耗時", round(max_time, 2), "秒"],
        ["最短耗時", round(min_time, 2), "秒"],
        ["", "", ""],
        ["📝 內容統計", "數值", "單位"],
        ["平均字數", int(avg_chars), "字元"],
    ]
    
    start_row = 5
    for row_idx, row_data in enumerate(summary_data, start=start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # 標題列樣式
            if row_data[0] in ["📈 整體統計", "🔧 爬蟲方法", "⏱️ 效能統計", "📝 內容統計"]:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.font = Font(bold=True, size=11)
            
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
    
    # 失敗項目列表
    fail_start = start_row + len(summary_data) + 2
    ws_summary.cell(row=fail_start, column=1).value = "❌ 失敗項目詳情"
    ws_summary.cell(row=fail_start, column=1).font = Font(bold=True, size=12, color="C00000")
    
    fail_headers = ["編號", "連結", "錯誤原因"]
    for col_idx, header in enumerate(fail_headers, 1):
        cell = ws_summary.cell(row=fail_start+1, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    fail_row = fail_start + 2
    for result in results:
        if not result.get('success'):
            url = result.get('url', '')
            id_num = result.get('id', '???')
            error = result.get('error', '未知錯誤')
            
            fail_data = [id_num, url, error]
            for col_idx, value in enumerate(fail_data, 1):
                cell = ws_summary.cell(row=fail_row, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            fail_row += 1
    
    # 成功案例展示（前10個）
    success_start = fail_row + 2
    ws_summary.cell(row=success_start, column=1).value = "✅ 成功案例展示（前10個）"
    ws_summary.cell(row=success_start, column=1).font = Font(bold=True, size=12, color="008000")
    
    success_headers = ["編號", "標題", "作者", "字數", "爬蟲方法"]
    for col_idx, header in enumerate(success_headers, 1):
        cell = ws_summary.cell(row=success_start+1, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    success_row = success_start + 2
    success_items = [r for r in results if r.get('success')][:10]
    for result in success_items:
        parsed_data = result.get('parsed_data', {})
        id_num = result.get('id', '???')
        title = parsed_data.get('title') or '無標題'
        author = parsed_data.get('author') or '未知'
        word_count = parsed_data.get('word_count') or 0
        method = result.get('rendering_method') or parsed_data.get('rendering_method', 'unknown')
        method_display = '動態' if method == 'playwright' else '靜態'
        
        success_data = [id_num, title, author, word_count, method_display]
        for col_idx, value in enumerate(success_data, 1):
            cell = ws_summary.cell(row=success_row, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        success_row += 1
    
    # 調整 Summary 列寬
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 60
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 12
    
    # 儲存檔案
    if not output_file:
        output_file = f"測試報告_{total}連結_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb.save(output_file)
    
    print("=" * 80)
    print("✅ Excel 報告生成成功！")
    print("=" * 80)
    print(f"📁 檔案名稱：{output_file}")
    print(f"📂 儲存位置：{os.path.abspath(output_file)}")
    print()
    print("📊 報告內容：")
    print(f"   • 分頁 1：詳細資料 (Master) - {len(results)} 筆資料 × 15 個欄位")
    print(f"   • 分頁 2：測試總結 (Summary) - 統計 + 失敗項目 + 成功案例")
    print()
    print(f"📈 統計摘要：")
    print(f"   • 成功率：{success_rate:.1f}% ({success_count}/{total})")
    print(f"   • 靜態爬蟲：{static_count} 個")
    print(f"   • 動態爬蟲：{dynamic_count} 個")
    print(f"   • 平均耗時：{avg_time:.2f} 秒")
    print("=" * 80)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法：python generate-excel-report.py <json檔案> [輸出檔名]")
        print("範例：python generate-excel-report.py results-120-raw.json")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(json_file):
        print(f"❌ 找不到檔案：{json_file}")
        sys.exit(1)
    
    generate_excel_report(json_file, output_file)

